from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill('solid', start_color='6366F1')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
CELL_FONT = Font(name='Arial', size=11)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
THIN = Side(border_style='thin', color='E5E7EB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 28

def style_data_cells(ws, start_row, end_row, n_cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=row, column=col)
            c.font = CELL_FONT
            c.border = BORDER
        ws.row_dimensions[row].height = 22

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============ CHARGES ============
ws = wb.create_sheet('Charges')
ws.append(['Date', 'Libellé', 'Catégorie', 'Montant', 'Payé par'])
charges_data = [
    (date(2026, 5, 1),  'Loyer',               'Logement',     850.00,  'Moi'),
    (date(2026, 5, 2),  'Courses Carrefour',   'Alimentation', 124.50,  'Femme'),
    (date(2026, 5, 5),  'Essence',             'Transport',     65.00,  'Moi'),
    (date(2026, 5, 10), 'Électricité EDF',     'Énergie',       95.00,  'Moi'),
    (date(2026, 5, 15), 'Abonnement Netflix',  'Loisirs',       15.99,  'Commun'),
]
for r in charges_data:
    ws.append(r)
style_header_row(ws, 1, 5)
style_data_cells(ws, 2, 6, 5)
set_widths(ws, [13, 28, 18, 14, 14])
for row in range(2, 7):
    ws.cell(row=row, column=1).number_format = 'YYYY-MM-DD'
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=2).alignment = LEFT
    ws.cell(row=row, column=3).alignment = LEFT
    ws.cell(row=row, column=4).number_format = '#,##0.00 €'
    ws.cell(row=row, column=4).alignment = RIGHT
    ws.cell(row=row, column=5).alignment = CENTER
ws.freeze_panes = 'A2'

# ============ REVENUS ============
ws = wb.create_sheet('Revenus')
ws.append(['Date', 'Libellé', 'Montant', 'Perçu par'])
revenus_data = [
    (date(2026, 5, 1), 'Salaire', 2400.00, 'Moi'),
    (date(2026, 5, 1), 'Salaire', 2100.00, 'Femme'),
]
for r in revenus_data:
    ws.append(r)
style_header_row(ws, 1, 4)
style_data_cells(ws, 2, 3, 4)
set_widths(ws, [13, 28, 14, 14])
for row in range(2, 4):
    ws.cell(row=row, column=1).number_format = 'YYYY-MM-DD'
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=2).alignment = LEFT
    ws.cell(row=row, column=3).number_format = '#,##0.00 €'
    ws.cell(row=row, column=3).alignment = RIGHT
    ws.cell(row=row, column=4).alignment = CENTER
ws.freeze_panes = 'A2'

# ============ CREDIT ============
ws = wb.create_sheet('Credit')
ws.append(['Date', 'Montant remboursé', 'Montant restant', 'Commentaire'])
credit_data = [
    (date(2026, 1, 15),    0.00, 145000.00, 'Solde initial du prêt'),
    (date(2026, 2, 15),  850.00, 144150.00, 'Mensualité février'),
    (date(2026, 3, 15),  850.00, 143300.00, 'Mensualité mars'),
    (date(2026, 4, 15),  850.00, 142450.00, 'Mensualité avril'),
]
for r in credit_data:
    ws.append(r)
style_header_row(ws, 1, 4)
style_data_cells(ws, 2, 5, 4)
set_widths(ws, [13, 18, 18, 30])
for row in range(2, 6):
    ws.cell(row=row, column=1).number_format = 'YYYY-MM-DD'
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=2).number_format = '#,##0.00 €'
    ws.cell(row=row, column=2).alignment = RIGHT
    ws.cell(row=row, column=3).number_format = '#,##0.00 €'
    ws.cell(row=row, column=3).alignment = RIGHT
    ws.cell(row=row, column=4).alignment = LEFT
ws.freeze_panes = 'A2'

# ============ NOTICE ============
ws = wb.create_sheet('À LIRE', 0)
ws['A1'] = '📌 Modèle Finance Perso — Mode d\'emploi'
ws['A1'].font = Font(name='Arial', bold=True, size=16, color='6366F1')

notice = [
    '',
    '1. Importe ce fichier dans Google Sheets :',
    '   • Va sur sheets.google.com → Fichier → Importer → Téléverser → choisis ce .xlsx',
    '   • Option d\'importation : "Remplacer la feuille de calcul"',
    '',
    '2. Vérifie que les onglets s\'appellent bien : "Charges", "Revenus", "Credit"',
    '   (sans accent sur Credit !)',
    '',
    '3. Tu peux SUPPRIMER cet onglet "À LIRE" après import — il ne sert qu\'à expliquer.',
    '',
    '4. Tu peux GARDER ou SUPPRIMER les lignes d\'exemples — comme tu veux.',
    '   ⚠️ NE SUPPRIME PAS les entêtes en ligne 1 de chaque onglet.',
    '',
    '5. Copie le SHEET_ID depuis l\'URL et colle-le dans js/config.js :',
    '   docs.google.com/spreadsheets/d/[CET_ID_ICI]/edit',
    '',
    '6. Partage la Sheet avec ta femme (bouton Partager en haut à droite).',
    '',
    '— — — Structure attendue par l\'app — — —',
    '',
    'Onglet "Charges" : Date | Libellé | Catégorie | Montant | Payé par',
    'Onglet "Revenus" : Date | Libellé | Montant | Perçu par',
    'Onglet "Credit"  : Date | Montant remboursé | Montant restant | Commentaire',
]

for i, line in enumerate(notice, start=2):
    ws.cell(row=i, column=1, value=line).font = Font(name='Arial', size=11)

ws.column_dimensions['A'].width = 90

wb.save('finance-perso-modele.xlsx')
print('OK: finance-perso-modele.xlsx créé')
